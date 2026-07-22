"""
reports.py
-----------
Export helpers for CSV, Excel (openpyxl), and PDF (reportlab) reports.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------
def to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode("utf-8")


# ---------------------------------------------------------------------------
# Excel (openpyxl)
# ---------------------------------------------------------------------------
def to_excel_bytes(df: pd.DataFrame, kpis: dict = None) -> bytes:
    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=13)
    title_font = Font(bold=True, size=16, color="1F4E78")

    ws_summary["B2"] = "Logistics Cost Analyzer — Summary Report"
    ws_summary["B2"].font = title_font
    ws_summary["B3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary["B3"].font = Font(italic=True, size=10, color="666666")

    if kpis:
        row = 5
        ws_summary.cell(row=row, column=2, value="Metric").font = header_font
        ws_summary.cell(row=row, column=2).fill = header_fill
        ws_summary.cell(row=row, column=3, value="Value").font = header_font
        ws_summary.cell(row=row, column=3).fill = header_fill
        row += 1
        for k, v in kpis.items():
            ws_summary.cell(row=row, column=2, value=k)
            ws_summary.cell(row=row, column=3, value=v)
            row += 1

    ws_summary.column_dimensions["A"].width = 3
    ws_summary.column_dimensions["B"].width = 32
    ws_summary.column_dimensions["C"].width = 22

    # --- Data sheet ---
    ws_data = wb.create_sheet("Shipment Data")
    ws_data.append(list(df.columns))
    for cell in ws_data[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    df_export = df.copy()
    if "date" in df_export.columns:
        df_export["date"] = df_export["date"].astype(str)

    for row_data in df_export.itertuples(index=False):
        ws_data.append(list(row_data))

    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws_data.iter_rows(min_row=1, max_row=ws_data.max_row):
        for cell in row:
            cell.border = border

    for i, col in enumerate(df.columns, start=1):
        max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
        ws_data.column_dimensions[get_column_letter(i)].width = min(max_len, 30)

    ws_data.freeze_panes = "A2"

    # --- Optional chart: monthly profit if available ---
    if "month" in df.columns and "profit" in df.columns:
        ws_chart = wb.create_sheet("Monthly Profit Chart")
        monthly = df.groupby("month", as_index=False)["profit"].sum().sort_values("month")
        ws_chart.append(["Month", "Profit"])
        for r in monthly.itertuples(index=False):
            ws_chart.append([r.month, r.profit])

        chart = BarChart()
        chart.title = "Monthly Profit"
        chart.y_axis.title = "Profit"
        chart.x_axis.title = "Month"
        data_ref = Reference(ws_chart, min_col=2, min_row=1, max_row=len(monthly) + 1)
        cats_ref = Reference(ws_chart, min_col=1, min_row=2, max_row=len(monthly) + 1)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws_chart.add_chart(chart, "D2")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF (reportlab)
# ---------------------------------------------------------------------------
def to_pdf_bytes(df: pd.DataFrame, kpis: dict, insights: list = None, max_rows: int = 60) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleStyle", parent=styles["Title"], textColor=colors.HexColor("#1F4E78"), fontSize=22
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], textColor=colors.HexColor("#666666"), fontSize=10
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"], textColor=colors.HexColor("#1F4E78"), spaceBefore=14
    )

    elements = []
    elements.append(Paragraph("Logistics Cost Analyzer", title_style))
    elements.append(Paragraph(f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    elements.append(Spacer(1, 16))

    # KPI table
    elements.append(Paragraph("Key Performance Indicators", section_style))
    kpi_items = list(kpis.items())
    kpi_rows = [["Metric", "Value", "Metric", "Value"]]
    for i in range(0, len(kpi_items), 2):
        left = kpi_items[i]
        right = kpi_items[i + 1] if i + 1 < len(kpi_items) else ("", "")
        kpi_rows.append([left[0], str(left[1]), right[0], str(right[1])])

    kpi_table = Table(kpi_rows, colWidths=[2.3 * inch, 1.7 * inch, 2.3 * inch, 1.7 * inch])
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 18))

    # Insights
    if insights:
        elements.append(Paragraph("Business Insights", section_style))
        for ins in insights:
            elements.append(Paragraph(f"<b>{ins['icon']} {ins['title']}</b> — {ins['detail']}", styles["Normal"]))
            elements.append(Spacer(1, 6))
        elements.append(Spacer(1, 12))

    # Data sample table
    elements.append(Paragraph(f"Shipment Data (showing up to {max_rows} rows)", section_style))
    display_cols = ["shipment_id", "date", "supplier", "route", "vehicle_type",
                     "total_cost", "revenue", "profit", "delivery_status"]
    display_cols = [c for c in display_cols if c in df.columns]
    sample = df[display_cols].head(max_rows).copy()
    if "date" in sample.columns:
        sample["date"] = sample["date"].astype(str)
    for c in ["total_cost", "revenue", "profit"]:
        if c in sample.columns:
            sample[c] = sample[c].map(lambda x: f"${x:,.2f}")

    table_data = [list(sample.columns)] + sample.values.tolist()
    col_widths = [1.1 * inch] * len(display_cols)
    data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    data_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D9D9D9")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(data_table)

    if len(df) > max_rows:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(
            f"<i>{len(df) - max_rows} additional rows not shown. Export to Excel or CSV for the full dataset.</i>",
            styles["Normal"]
        ))

    doc.build(elements)
    return buffer.getvalue()
